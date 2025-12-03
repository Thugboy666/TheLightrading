import io
import logging
import os
import uuid
from datetime import datetime, timedelta, timezone
from logging.handlers import RotatingFileHandler
from pathlib import Path
from typing import Any, Dict, List, Optional

import bcrypt
import hashlib
import json
import httpx
from aiohttp import web
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from .db import (
    clear_discount_rules_for_offer_segment,
    create_session_with_expiry,
    create_user,
    add_promo_points,
    delete_client,
    delete_product,
    get_client_by_email,
    find_client_by_email_or_piva,
    get_daily_offer,
    get_product_by_sku,
    get_promo_summary,
    get_session,
    get_user_by_email,
    get_user_by_id,
    init_db,
    list_orders,
    insert_discount_rule,
    link_client_to_user_by_email,
    list_clients,
    list_discount_rules,
    list_products,
    save_client,
    save_import_metadata,
    set_meta_value,
    save_daily_offer,
    update_user_password,
    upsert_product,
    delete_daily_offer,
    get_meta_value,
)

# ================== CONFIG BASE ==================

BASE_DIR = Path(__file__).resolve().parent.parent

LOG_DIR = BASE_DIR / "logs"
LOG_DIR.mkdir(exist_ok=True)

LOG_FILE = LOG_DIR / "api.log"

logger = logging.getLogger("thelight24.api")
logger.setLevel(logging.INFO)

# Evita handler duplicati se il modulo viene ricaricato
if not logger.handlers:
    handler = RotatingFileHandler(
        LOG_FILE,
        maxBytes=2 * 1024 * 1024,  # 2 MB
        backupCount=3,
        encoding="utf-8",
    )
    formatter = logging.Formatter(
        "%(asctime)s [%(levelname)s] %(name)s - %(message)s"
    )
    handler.setFormatter(formatter)
    logger.addHandler(handler)

    # Log anche su stdout (utile in fase debug)
    console = logging.StreamHandler()
    console.setFormatter(formatter)
    logger.addHandler(console)

logger.info("TheLight24 API avviata (server.py caricato)")


def log_event(event: str, **extra: Any) -> None:
    """
    Logga un evento applicativo strutturato.
    Esempio: log_event("user_login", email=email, tier=tier)
    """
    payload = {"event": event}
    payload.update(extra)
    try:
        logger.info(json.dumps(payload, ensure_ascii=False))
    except Exception:
        # Fallback se qualcosa in extra non è serializzabile
        logger.info("%s | extra=%r", event, extra)


def _resolve_ui_index() -> Path:
    """Return the index.html path (overridable via THELIGHT_UI_INDEX)."""

    candidate = os.environ.get("THELIGHT_UI_INDEX")
    if candidate:
        ui_path = Path(candidate).expanduser().resolve()
    else:
        ui_path = BASE_DIR / "gui" / "index.html"

    if not ui_path.is_file():
        raise FileNotFoundError(f"GUI index non trovato: {ui_path}")

    return ui_path


UI_INDEX = _resolve_ui_index()

# Initialize persistence layer
init_db()

# backend LLM locale (Termux / llama.cpp / phi ecc.)
#   - LLM_BACKEND_URL ha priorità e può puntare già all'endpoint completo
#   - altrimenti usiamo THELIGHT_LLM_BASE_URL / completions di default
LLM_BACKEND_URL = os.environ.get("LLM_BACKEND_URL", "http://127.0.0.1:8081/completion")


# ================== FALLBACK UTILS ==================

def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")


def hash_password_legacy(password: str) -> str:
    return hashlib.sha256(password.encode()).hexdigest()


def verify_password(plain_password: str, stored_hash: str) -> bool:
    if not stored_hash:
        return False
    # Bcrypt hash
    try:
        if stored_hash.startswith("$2") and bcrypt.checkpw(
            plain_password.encode("utf-8"), stored_hash.encode("utf-8")
        ):
            return True
    except Exception:
        # fallback to legacy checks
        pass
    # Legacy sha256 hash
    if stored_hash == hash_password_legacy(plain_password):
        return True
    # Plain text fallback (vecchi import non hashati)
    if stored_hash == plain_password:
        return True
    return False


def normalize_header(value: str) -> str:
    if not value:
        return ""
    return (
        str(value)
        .strip()
        .lower()
        .replace("à", "a")
        .replace("è", "e")
        .replace("é", "e")
        .replace("ì", "i")
        .replace("ò", "o")
        .replace("ù", "u")
    )


def safe_float(value, default=0.0):
    if value is None:
        return default
    s = str(value).strip()
    if not s:
        return default
    s = s.replace(" ", "").replace("€", "")
    if s.count(",") == 1 and s.count(".") == 0:
        s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return default


def safe_int(value, default=0):
    if value is None:
        return default
    s = str(value).strip()
    if not s:
        return default

    s_upper = s.upper()
    if s_upper in {"N", "NA", "N/A", "ND", "-", "NON DISPONIBILE"}:
        return default
    s_clean = s.replace(" ", "")
    if s_clean.replace(".", "").replace(",", "").isdigit():
        s_clean = s_clean.replace(".", "").replace(",", "")
    try:
        return int(s_clean)
    except ValueError:
        return default


def discount_rules_to_configs() -> list[dict]:
    configs: dict[str, dict] = {}
    for row in list_discount_rules():
        cfg = configs.setdefault(row["offer_id"], {"id": row["offer_id"], "rules": {}})
        seg_rules = cfg["rules"].setdefault(row["segment"], [])
        seg_rules.append(
            {
                "min": row["min_amount"],
                "max": row["max_amount"] if row["max_amount"] is not None else float("inf"),
                "discount": row["discount_percent"],
                "valid_until": row.get("valid_until"),
            }
        )
    return list(configs.values())


def pick_price_for_segment(product: dict, segment: str, fallback_base: float) -> float:
    if segment == "distributore":
        if product.get("price_distributore") is not None:
            return float(product.get("price_distributore"))
        if product.get("price_dist") is not None:
            return float(product.get("price_dist"))
    if segment == "rivenditore":
        if product.get("price_rivenditore") is not None:
            return float(product.get("price_rivenditore"))
        if product.get("price_riv") is not None:
            return float(product.get("price_riv"))
    if segment == "rivenditore10":
        if product.get("price_rivenditore10") is not None:
            return float(product.get("price_rivenditore10"))
        if product.get("price_riv10") is not None:
            return float(product.get("price_riv10"))
    return float(product.get("base_price") or fallback_base or 0)


def compute_price_with_discounts(product: dict, customer_segment: str, quantity: int) -> dict:
    base_price = pick_price_for_segment(product, customer_segment, product.get("base_price", 0))
    amount = base_price * max(1, quantity)
    offer_id = None
    extra = product.get("extra") or {}
    if isinstance(extra, dict):
        offer_id = extra.get("offer_id")

    discount_value = 0.0
    if offer_id:
        for cfg in discount_rules_to_configs():
            if cfg.get("id") != offer_id:
                continue
            rules = cfg.get("rules", {}).get(customer_segment, [])
            for rule in rules:
                max_amount = rule.get("max")
                if max_amount is None:
                    max_amount = float("inf")
                if amount >= float(rule.get("min", 0)) and amount <= float(max_amount):
                    discount_value = (rule.get("discount", 0) or 0) / 100.0
                    break
            if discount_value:
                break

    final_amount = amount * (1 - discount_value)
    return {
        "sku": product.get("sku"),
        "base_price": base_price,
        "segment": customer_segment,
        "quantity": quantity,
        "price": final_amount,
        "discount_applied": discount_value * 100,
    }


def client_payload_from_record(client: Optional[dict]) -> Optional[dict]:
    if not client:
        return None

    return {
        "id": client.get("id"),
        "ragione_sociale": client.get("ragione_sociale"),
        "piva": client.get("piva"),
        "email": client.get("email"),
        "telefono": client.get("telefono"),
        "phone": client.get("telefono"),
        "listino": client.get("listino"),
        "stato": client.get("stato"),
        "promo_enabled": bool(int(client.get("promo_enabled") or 0)),
        "promo_points": int(client.get("promo_points") or 0),
        "promo_ticket_code": client.get("promo_ticket_code"),
        "user_id": client.get("user_id"),
        "user_tier": client.get("user_tier"),
    }


# ================== HANDLER FRONTEND ==================

async def ui_index(request: web.Request) -> web.Response:
    """Serve la GUI 3D (index.html)."""
    return web.FileResponse(path=UI_INDEX)


# ================== SYSTEM ==================

async def health(request: web.Request) -> web.Response:
    data = {
        "status": "ok",
        "project": "TheLight24 v7",
        "env": "dev",
        "time": now_iso(),
    }
    return web.json_response(data)


# ================== AUTH (STUB) ==================

async def auth_register(request: web.Request) -> web.Response:
    body = await request.json()
    if not body.get("accept_terms"):
        return web.json_response({"status": "ko", "error": "terms"})

    email = body.get("email")
    if not email:
        return web.json_response({"status": "ko", "error": "missing_email"})

    if get_user_by_email(email):
        return web.json_response({"status": "ko", "error": "exists"})

    name = body.get("name") or email
    temp_password = body.get("password") or "changeme"
    password_hash = hash_password(temp_password)
    tier = "rivenditore10"
    create_user(
        email=email,
        password_hash=password_hash,
        name=name,
        tier=tier,
        piva=body.get("piva") or "",
        phone=body.get("phone") or "",
    )
    link_client_to_user_by_email(email, create_missing_client=True, default_listino=tier)
    log_event(
        "user_registered",
        email=email,
        tier=tier,
        piva=body.get("piva") or "",
    )
    return web.json_response({"status": "ok", "tier": tier})


async def auth_login(request: web.Request) -> web.Response:
    body = await request.json()
    email = body.get("email", "").strip()
    password = body.get("password", "")
    remember = bool(body.get("remember", False))

    ADMIN_EMAIL = "god@local"
    ADMIN_PASS = "OrmaNet!2025$Light"
    if email == ADMIN_EMAIL and password == ADMIN_PASS:
        log_event("admin_login", email=email, tier="distributore")
        return web.json_response(
            {
                "status": "ok",
                "name": "GOD ADMIN",
                "tier": "distributore",
                "token": None,
                "is_admin": True,
            }
        )

    user = get_user_by_email(email)
    if not user:
        log_event("user_login_failed", email=email)
        return web.json_response({"status": "ko"})

    if not verify_password(password, user.get("password_hash")):
        log_event("user_login_failed", email=email)
        return web.json_response({"status": "ko"})

    expires_delta = timedelta(days=30) if remember else timedelta(hours=24)
    token = create_session_with_expiry(user_id=user["id"], expires_delta=expires_delta)
    link_client_to_user_by_email(email)
    log_event(
        "user_login",
        email=email,
        tier=user.get("tier", "rivenditore10"),
        token=token,
    )
    return web.json_response(
        {
            "status": "ok",
            "name": user.get("name") or user.get("email"),
            "tier": user.get("tier", "rivenditore10"),
            "token": token,
            "is_admin": bool(user.get("is_admin")),
        }
    )


async def auth_validate_session(request: web.Request) -> web.Response:
    token = request.headers.get("Authorization", "").replace("Bearer", "").strip()
    if not token:
        return web.json_response({"valid": False})
    session = get_session(token)
    if not session:
        return web.json_response({"valid": False})
    return web.json_response({"valid": True, "token": token})


async def auth_me(request: web.Request) -> web.Response:
    token = request.headers.get("Authorization", "").replace("Bearer", "").strip()
    if not token:
        return web.json_response({"status": "error", "message": "Missing token"}, status=401)

    session = get_session(token)
    if not session:
        return web.json_response({"status": "error", "message": "Token non valido"}, status=401)

    user = get_user_by_id(session.get("user_id"))
    if not user:
        return web.json_response({"status": "error", "message": "Utente non trovato"}, status=404)

    client = link_client_to_user_by_email(user.get("email")) or get_client_by_email(
        user.get("email")
    )

    return web.json_response(
        {
            "user": {
                "email": user.get("email"),
                "name": user.get("name"),
                "tier": user.get("tier"),
                "id": user.get("id"),
                "piva": user.get("piva"),
                "phone": user.get("phone"),
                "is_admin": bool(user.get("is_admin")),
            },
            "client": client_payload_from_record(client),
            # Chiavi legacy mantenute per compatibilità con il frontend esistente
            "email": user.get("email"),
            "name": user.get("name"),
            "tier": user.get("tier"),
            "id": user.get("id"),
            "is_admin": bool(user.get("is_admin")),
        }
    )


async def auth_change_password(request: web.Request) -> web.Response:
    token = request.headers.get("Authorization", "").replace("Bearer", "").strip()
    if not token:
        return web.json_response({"status": "error", "message": "Token mancante"}, status=401)

    session = get_session(token)
    if not session:
        return web.json_response({"status": "error", "message": "Sessione non valida"}, status=401)

    user = get_user_by_id(session.get("user_id"))
    if not user:
        return web.json_response({"status": "error", "message": "Utente non trovato"}, status=404)

    body = await request.json()
    current_password = (body.get("current_password") or "").strip()
    new_password = (body.get("new_password") or "").strip()

    if not current_password or not new_password:
        return web.json_response(
            {"status": "error", "message": "Compila tutti i campi"}, status=400
        )

    if len(new_password) < 8:
        return web.json_response(
            {"status": "error", "message": "Nuova password troppo corta"}, status=400
        )

    if not verify_password(current_password, user.get("password_hash")):
        return web.json_response(
            {"status": "error", "message": "Password attuale non corretta."},
            status=400,
        )

    new_hash = hash_password(new_password)
    update_user_password(user_id=user["id"], new_password_hash=new_hash)
    log_event("user_password_changed", email=user.get("email"))
    return web.json_response({"status": "ok"})


# ================== ECOMMERCE ==================

async def pricing(request: web.Request) -> web.Response:
    body = await request.json()
    sku = body.get("sku")
    segment = body.get("customer_segment", "ospite")
    qty = int(body.get("quantity", 1))
    product = get_product_by_sku(sku) if sku else None
    if not product:
        product = {
            "sku": sku,
            "base_price": body.get("base_price", 0),
            "extra": {"offer_id": body.get("offer_id")},
        }
    result = compute_price_with_discounts(product, segment, qty)
    return web.json_response(result)


async def order_draft(request: web.Request) -> web.Response:
    body = await request.json()
    log_event(
        "order_draft",
        user_tier=body.get("user_tier"),
        user_name=body.get("user_name"),
        items=len(body.get("items") or []),
    )
    return web.json_response({"status": "ok", "received": body})


async def product_update(request: web.Request) -> web.Response:
    body = await request.json()
    pricing = body.get("pricing", {})
    prices = pricing.get("prices", {}) if isinstance(pricing, dict) else {}
    product = {
        "sku": body.get("sku") or str(uuid.uuid4()),
        "name": body.get("name"),
        "image_hd": body.get("image_hd"),
        "image_thumb": body.get("image_thumb"),
        "gallery": body.get("gallery") or [],
        "description_html": body.get("description_html"),
        "base_price": pricing.get("base_price"),
        "unit": pricing.get("unit"),
        "markup_riv10": pricing.get("markup_riv10"),
        "markup_riv": pricing.get("markup_riv"),
        "markup_dist": pricing.get("markup_dist"),
        "price_riv10": prices.get("rivenditore10"),
        "price_riv": prices.get("rivenditore"),
        "price_dist": prices.get("distributore"),
        "extra": body.get("extra") or {},
    }
    saved = upsert_product(product)
    return web.json_response({"status": "ok", "product": saved})


async def account_orders(request: web.Request) -> web.Response:
    token = request.headers.get("Authorization", "").replace("Bearer", "").strip()
    if not token:
        return web.json_response({"status": "error", "message": "Token mancante"}, status=401)

    session = get_session(token)
    if not session:
        return web.json_response({"status": "error", "message": "Sessione non valida"}, status=401)

    user = get_user_by_id(session.get("user_id"))
    if not user:
        return web.json_response({"status": "error", "message": "Utente non trovato"}, status=404)

    client = get_client_by_email(user.get("email"))
    company_name = None
    if client:
        company_name = client.get("ragione_sociale") or client.get("name")

    params = request.rel_url.query
    status_param = params.get("status") or None
    cause_param = params.get("cause") or None
    date_from = params.get("date_from") or None
    date_to = params.get("date_to") or None

    orders = list_orders(
        customer_email=user.get("email"),
        customer_name=company_name,
        status=status_param,
        cause=cause_param,
        date_from=date_from,
        date_to=date_to,
        include_all=bool(user.get("is_admin")),
    )

    return web.json_response({"orders": orders})


async def admin_clients_all(request: web.Request) -> web.Response:
    clients: list[dict] = []
    for client in list_clients():
        payload = client_payload_from_record(client) or {}
        payload.update(
            {
                "note": client.get("note"),
                "promo_last_update": client.get("promo_last_update"),
                "created_at": client.get("created_at"),
                "updated_at": client.get("updated_at"),
            }
        )
        # Conserva i nomi legacy per compatibilità con il frontend admin
        payload.setdefault("telefono", client.get("telefono"))
        payload.setdefault("email", client.get("email"))
        clients.append(payload)

    return web.json_response({"clients": clients})


async def admin_clients_save(request: web.Request) -> web.Response:
    body = await request.json()
    saved = save_client(body)
    linked = link_client_to_user_by_email(saved.get("email"))
    if linked:
        saved["user_id"] = linked.get("user_id")
    log_event(
        "client_save",
        id=saved.get("id"),
        ragione_sociale=saved.get("ragione_sociale") or saved.get("name"),
        listino=saved.get("listino"),
    )
    return web.json_response(saved)


async def admin_clients_delete(request: web.Request) -> web.Response:
    body = await request.json()
    if body.get("id"):
        delete_id = int(body["id"])
        delete_client(delete_id)
        log_event("client_delete", id=delete_id)
    return web.json_response({"status": "ok"})


async def admin_clients_import_promo(request: web.Request) -> web.Response:
    """
    Importa anagrafiche aderenti alla promo da file Excel (preferibilmente .xlsx)
    Formato colonne SENZA intestazione:
      0: tipo (es. 'azienda' / 'privato')
      1: ragione sociale / nome
      2: P.IVA
      3: SDI / PEC
      4: regime fiscale
      5: indirizzo completo
      6: email
      7: telefono
      8: come ci hai conosciuti
      9: metodo pagamento preferito
     10: listino (distributore / rivenditore / rivenditore10)
    """

    data = await request.post()
    upfile = data.get("file")
    if not upfile:
        logger.warning("Import promo: missing file upload")
        return web.json_response(
            {"status": "error", "reason": "missing_file"},
            status=400,
        )

    filename = getattr(upfile, "filename", "") or ""
    if not filename or "." not in filename:
        logger.warning("Import promo: invalid filename")
        return web.json_response(
            {"status": "error", "reason": "invalid_filename"},
            status=400,
        )

    ext = Path(filename).suffix.lower()
    if ext not in {".xls", ".xlsx"}:
        logger.warning("Import promo: unsupported extension %s", ext)
        return web.json_response(
            {"status": "error", "reason": "unsupported_extension"},
            status=400,
        )

    if ext == ".xls":
        logger.warning("Import promo: xls received, ask for xlsx")
        return web.json_response(
            {
                "status": "error",
                "reason": "xls_not_supported",
                "message": "Salva il file in formato .xlsx e riprova.",
            },
            status=400,
        )

    try:
        if hasattr(upfile, "file"):
            file_bytes = upfile.file.read()
        else:
            file_bytes = upfile.read()
    except Exception as exc:  # noqa: BLE001
        logger.exception("Import promo: unable to read file %s: %s", filename, exc)
        return web.json_response(
            {"status": "error", "reason": "read_error"},
            status=400,
        )

    if not file_bytes:
        logger.warning("Import promo: uploaded file %s is empty", filename)
        return web.json_response(
            {"status": "error", "reason": "empty_file"},
            status=400,
        )

    try:
        wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    except InvalidFileException:
        logger.exception("Import promo: invalid Excel file %s", filename)
        return web.json_response(
            {"status": "error", "reason": "invalid_excel"},
            status=400,
        )
    except Exception as exc:  # noqa: BLE001
        logger.exception("Import promo: error loading workbook %s: %s", filename, exc)
        return web.json_response(
            {"status": "error", "reason": "invalid_excel"},
            status=400,
        )

    if not wb.sheetnames:
        logger.error("Import promo: workbook %s has no sheets", filename)
        return web.json_response(
            {"status": "error", "reason": "no_sheets"},
            status=400,
        )

    sheet = wb[wb.sheetnames[0]]

    try:
        rows = []
        for row in sheet.iter_rows(values_only=True):
            if not any(row):
                continue
            rows.append(row)
    except Exception as exc:  # noqa: BLE001
        logger.exception("Import promo: error reading rows from %s: %s", filename, exc)
        return web.json_response(
            {"status": "error", "reason": "invalid_excel"},
            status=400,
        )

    if not rows:
        logger.warning("Import promo: workbook %s has no data rows", filename)
        return web.json_response(
            {"status": "error", "reason": "no_rows"},
            status=400,
        )

    def _clean(value: Any) -> Optional[str]:
        if value is None:
            return None
        if isinstance(value, str):
            return value.strip()
        return str(value).strip()

    def _valid_email(value: Optional[str]) -> bool:
        if not value:
            return False
        return "@" in value and "." in value.split("@")[-1]

    imported = 0
    updated = 0
    processed = 0
    errors: list[dict[str, Any]] = []

    logger.info("Import promo: processing file %s (%d rows)", filename, len(rows))

    for idx, row in enumerate(rows, start=1):
        processed += 1
        try:
            cols = list(row) + [None] * (11 - len(row))
            (
                tipo,
                ragione_sociale,
                piva,
                sdi,
                regime_fiscale,
                indirizzo,
                email,
                telefono,
                _source,
                _payment_pref,
                listino,
            ) = cols[:11]

            ragione_sociale = _clean(ragione_sociale) or None
            piva = _clean(piva) or None
            email = _clean(email) or None
            telefono = _clean(telefono) or None
            listino = (_clean(listino) or "rivenditore10")

            if not email and not piva:
                errors.append({"row": idx, "error": "missing_email_and_piva"})
                continue

            if email and not _valid_email(email):
                errors.append({"row": idx, "error": "invalid_email", "email": email})
                continue

            client_data = {
                "id": None,
                "ragione_sociale": ragione_sociale,
                "piva": piva,
                "email": email,
                "telefono": telefono,
                "listino": listino,
                "stato": "attivo",
                "note": "",
                "promo_enabled": 1,
                "promo_points": 0,
                "promo_ticket_code": None,
            }

            existing = find_client_by_email_or_piva(email, piva)
            if existing:
                client_data["id"] = existing.get("id")
                save_client(client_data)
                updated += 1
            else:
                save_client(client_data)
                imported += 1

            link_client_to_user_by_email(email)
        except Exception as exc:  # noqa: BLE001
            logger.exception("Import promo: error processing row %s", idx)
            errors.append({"row": idx, "error": "exception", "detail": str(exc)})

    return web.json_response(
        {
            "status": "ok",
            "processed": processed,
            "imported": imported,
            "updated": updated,
            "errors": errors,
        }
    )


PROMO_POINTS = {
    "FOLLOW_SOCIAL": 5,
    "ADD_BROADCAST": 50,
    "ORDER_REMAN": 100,
    "REACH_AVG_REVENUE": 200,
    "UPSELL_REVENUE": 500,
    "BRING_NEW_COMPANY": 1000,
}


async def admin_promo_add_points(request: web.Request) -> web.Response:
    body = await request.json()
    client_id = body.get("client_id")
    action_code = body.get("action_code")
    if not client_id or action_code not in PROMO_POINTS:
        return web.json_response({"status": "ko", "error": "invalid_payload"}, status=400)

    updated_client = add_promo_points(int(client_id), action_code, PROMO_POINTS[action_code])
    if not updated_client:
        return web.json_response({"status": "ko", "error": "client_not_found"}, status=404)

    summary = get_promo_summary(int(client_id))
    return web.json_response({"status": "ok", "client": updated_client, "summary": summary})


async def admin_promo_summary(request: web.Request) -> web.Response:
    client_id = request.query.get("client_id")
    if not client_id:
        return web.json_response({"status": "ko", "error": "missing_client_id"}, status=400)
    summary = get_promo_summary(int(client_id))
    if not summary:
        return web.json_response({"status": "ko", "error": "client_not_found"}, status=404)
    return web.json_response(summary)


async def admin_offers_all(request: web.Request) -> web.Response:
    return web.json_response({"configs": discount_rules_to_configs()})


async def admin_offers_save(request: web.Request) -> web.Response:
    body = await request.json()
    offer_id = body.get("id")
    rules = body.get("rules") or {}
    if not offer_id:
        return web.json_response({"status": "ko", "error": "missing_offer_id"}, status=400)

    for segment, items in rules.items():
        clear_discount_rules_for_offer_segment(offer_id, segment)
        for rule in items:
            insert_discount_rule(
                offer_id=offer_id,
                segment=segment,
                min_amount=float(rule.get("min", 0)),
                max_amount=None if rule.get("max") in (None, "", float("inf")) else float(rule.get("max")),
                discount_percent=float(rule.get("discount", 0)),
                valid_until=rule.get("valid_until"),
            )
    return web.json_response({"status": "ok", "config": body})


async def admin_products_all(request: web.Request) -> web.Response:
    raw_products = list_products()
    products = []
    for r in raw_products:
        gallery = r.get("gallery") if isinstance(r, dict) else r.get("gallery", [])
        if isinstance(gallery, str):
            try:
                gallery = json.loads(gallery)
            except Exception:
                gallery = []

        products.append(
            {
                "sku": r.get("sku"),
                "codice": r.get("codice"),
                "name": r.get("name"),
                "desc_html": r.get("desc_html") or r.get("description_html"),
                "image_hd": r.get("image_hd"),
                "image_thumb": r.get("image_thumb"),
                "gallery": gallery or [],
                "base_price": r.get("base_price"),
                "unit": r.get("unit"),
                "markup_riv10": r.get("markup_riv10"),
                "markup_riv": r.get("markup_riv"),
                "markup_dist": r.get("markup_dist"),
                "price_distributore": r.get("price_distributore")
                or r.get("price_dist"),
                "price_rivenditore": r.get("price_rivenditore") or r.get("price_riv"),
                "price_rivenditore10": r.get("price_rivenditore10")
                or r.get("price_riv10"),
                "qty_stock": r.get("qty_stock"),
                "discount_dist_percent": r.get("discount_dist_percent"),
                "discount_riv_percent": r.get("discount_riv_percent"),
                "discount_riv10_percent": r.get("discount_riv10_percent"),
                "status": r.get("status"),
            }
        )

    products.sort(key=lambda p: (p.get("sku") or ""))
    return web.json_response({"products": products})


async def admin_price_list_import(request: web.Request) -> web.Response:
    reader = await request.multipart()
    file_part = None

    async for part in reader:
        if part.name in ("file", "price_list_file", "price_list"):
            file_part = part
            break

    if file_part is None:
        return web.json_response({"error": "Nessun file di listino fornito."}, status=400)

    file_data = await file_part.read()
    if not file_data:
        return web.json_response({"error": "File listino vuoto."}, status=400)

    wb = load_workbook(io.BytesIO(file_data), data_only=True)
    ws = wb.active

    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    header_map: dict[str, int] = {}
    for idx, header in enumerate(header_row):
        key = normalize_header(header)
        if not key:
            continue
        header_map[key] = idx

    def get_cell(row, *candidates, default=None):
        for cand in candidates:
            cand_norm = normalize_header(cand)
            idx = header_map.get(cand_norm)
            if idx is not None and idx < len(row):
                return row[idx]
        return default

    inserted = 0
    updated = 0
    skipped = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row is None:
            continue

        codice = get_cell(row, "codice")
        descrizione = get_cell(
            row,
            "descrizione",
            "descrizione articolo",
            "descrizione_articolo",
            "nome",
            "nome articolo",
        )
        prezzo_distributore = get_cell(row, "prezzo_distributore", "prezzo distributore")
        prezzo_rivenditore = get_cell(row, "prezzo_rivenditore", "prezzo rivenditore")
        prezzo_rivenditore10 = get_cell(
            row,
            "prezzo_rivenditore10",
            "prezzo rivenditore10",
            "prezzo_rivenditore_10",
            "prezzo rivenditore 10",
        )
        qty_stock = get_cell(
            row,
            "quantità_stock",
            "quantita_stock",
            "quantita stock",
            "quantità stock",
            "qta",
            "giacenza",
        )
        status = get_cell(row, "status", "stato")

        if codice is not None:
            codice = str(codice).strip()
        if descrizione is not None:
            descrizione = str(descrizione).strip()

        if not codice and not descrizione:
            skipped += 1
            continue

        sku = codice or (descrizione[:20] if descrizione else "")

        base_price = safe_float(prezzo_rivenditore)
        price_dist = safe_float(prezzo_distributore)
        price_riv = safe_float(prezzo_rivenditore)
        price_riv10 = safe_float(prezzo_rivenditore10)
        qty = safe_int(qty_stock, 0)
        status_norm = (str(status).strip().upper() if status is not None else "N")

        product_data = {
            "sku": sku,
            "codice": codice or sku,
            "name": descrizione or sku,
            "desc_html": descrizione or "",
            "image_hd": None,
            "image_thumb": None,
            "gallery": [],
            "base_price": base_price,
            "unit": "pz",
            "markup_riv10": None,
            "markup_riv": None,
            "markup_dist": None,
            "price_distributore": price_dist,
            "price_rivenditore": price_riv,
            "price_rivenditore10": price_riv10,
            "qty_stock": qty,
            "discount_dist_percent": 0.0,
            "discount_riv_percent": 0.0,
            "discount_riv10_percent": 0.0,
            "status": "attivo" if status_norm == "S" else "non_disponibile",
        }

        existing = get_product_by_sku(sku)
        upsert_product(product_data)
        if existing:
            updated += 1
        else:
            inserted += 1

    total_processed = inserted + updated
    save_import_metadata(file_part.filename or "listino.xls", total_processed)
    log_event(
        "price_list_import",
        filename=file_part.filename or "listino.xls",
        imported_count=total_processed,
    )

    products = list_products()
    products.sort(key=lambda p: (p.get("sku") or ""))
    last_import_at = now_iso()
    set_meta_value("price_list_last_import_at", last_import_at)

    return web.json_response(
        {
            "inserted": inserted,
            "updated": updated,
            "skipped": skipped,
            "products": products,
            "last_import_at": last_import_at,
        }
    )


async def admin_price_list_status(request: web.Request) -> web.Response:
    db = request.app["db"]
    row = await db.fetch_one(
        "SELECT value FROM meta WHERE key = :key",
        {"key": "price_list_last_import_at"},
    )
    last_import_at = row["value"] if row else None
    return web.json_response({"last_import_at": last_import_at})


async def admin_product_save(request: web.Request) -> web.Response:
    body = await request.json()
    product = {
        "sku": body.get("sku") or str(uuid.uuid4()),
        "name": body.get("name"),
        "codice": body.get("codice"),
        "image_hd": body.get("image_hd"),
        "image_thumb": body.get("image_thumb"),
        "gallery": body.get("gallery") or [],
        "description_html": body.get("description_html") or body.get("desc_html"),
        "base_price": body.get("base_price"),
        "unit": body.get("unit"),
        "markup_riv10": body.get("markup_riv10"),
        "markup_riv": body.get("markup_riv"),
        "markup_dist": body.get("markup_dist"),
        "price_distributore": body.get("price_distributore"),
        "price_rivenditore": body.get("price_rivenditore"),
        "price_rivenditore10": body.get("price_rivenditore10"),
        "qty_stock": body.get("qty_stock", 0),
        "discount_dist_percent": body.get("discount_dist_percent", 0),
        "discount_riv_percent": body.get("discount_riv_percent", 0),
        "discount_riv10_percent": body.get("discount_riv10_percent", 0),
        "status": body.get("status") or "attivo",
    }
    saved = upsert_product(product)
    return web.json_response(saved)


async def admin_product_delete(request: web.Request) -> web.Response:
    sku = request.match_info.get("sku")
    if not sku:
        return web.json_response({"error": "missing_sku"}, status=400)
    deleted = delete_product(sku)
    if not deleted:
        return web.json_response({"status": "not_found"}, status=404)
    return web.json_response({"status": "ok"})


def _enrich_daily_offer(record: Optional[dict]) -> Optional[dict]:
    if not record:
        return None
    product = get_product_by_sku(record.get("sku"))
    if product:
        record["product_name"] = product.get("name")
    return record


async def admin_daily_offer_get(request: web.Request) -> web.Response:
    offer = _enrich_daily_offer(get_daily_offer())
    if not offer:
        return web.json_response({})
    return web.json_response(offer)


async def admin_daily_offer_save(request: web.Request) -> web.Response:
    body = await request.json()
    min_qty = safe_int(body.get("min_qty", 1), 1)
    if min_qty < 1:
        min_qty = 1
    body["min_qty"] = min_qty
    product_url = (body.get("product_url") or "").strip() or None
    body["product_url"] = product_url
    offer = save_daily_offer(body)
    enriched = _enrich_daily_offer(offer)
    return web.json_response(enriched)


async def admin_daily_offer_delete(request: web.Request) -> web.Response:
    delete_daily_offer()
    return web.json_response({"status": "ok"})


async def public_daily_offer(request: web.Request) -> web.Response:
    offer = get_daily_offer()
    if not offer or not offer.get("active"):
        return web.json_response({"active": False})

    now = datetime.now(timezone.utc)
    start_at = offer.get("start_at")
    end_at = offer.get("end_at")

    def parse_dt(value: Optional[str]) -> Optional[datetime]:
        if not value:
            return None
        try:
            return datetime.fromisoformat(value.replace("Z", "+00:00"))
        except Exception:
            return None

    start_dt = parse_dt(start_at)
    end_dt = parse_dt(end_at)
    if start_dt and now < start_dt:
        return web.json_response({"active": False})
    if end_dt and now > end_dt:
        return web.json_response({"active": False})

    product = get_product_by_sku(offer.get("sku"))
    response = {
        "active": True,
        "sku": offer.get("sku"),
        "product_name": product.get("name") if product else offer.get("product_name"),
        "coupon_code": offer.get("coupon_code"),
        "start_at": start_at,
        "end_at": end_at,
        "discount_dist_percent": offer.get("discount_dist_percent", 0),
        "discount_riv_percent": offer.get("discount_riv_percent", 0),
        "discount_riv10_percent": offer.get("discount_riv10_percent", 0),
        "min_qty": offer.get("min_qty", 1),
        "product_url": offer.get("product_url"),
    }
    return web.json_response(response)


# ================== LLM LOCALE ==================

async def llm_complete(request: web.Request) -> web.Response:
    """
    Endpoint interno TheLight24 (non usato dalla GUI attuale)
    POST /llm/complete
    """
    body = await request.json()
    prompt = body.get("prompt", "")
    max_tokens = int(body.get("max_tokens", 128))

    # stub finché non colleghi un client reale
    result = {
        "completion": f"[LLM stub interno] Richiesta: {prompt[:80]}...",
        "max_tokens": max_tokens,
    }
    return web.json_response(result)


async def llm_chat(request: web.Request) -> web.Response:
    """
    Endpoint usato dalla GUI:
    LLM_URL = '/api/llm/chat'
    Proxi verso il backend 127.0.0.1:8081/completion
    Normalizziamo SEMPRE in: {"content": "..."}
    """
    payload = await request.json()

    try:
        async with httpx.AsyncClient(timeout=120.0) as client:
            r = await client.post(LLM_BACKEND_URL, json=payload)
    except Exception as e:  # noqa: BLE001
        return web.json_response(
            {"error": f"impossibile contattare backend LLM: {e}"},
            status=502,
        )

    # Proviamo a fare parse JSON, altrimenti teniamo il testo puro
    text_body = r.text
    try:
        raw = r.json()
    except Exception:  # noqa: BLE001
        raw = None

    content = ""

    if isinstance(raw, dict):
        # 1) Caso classico llama.cpp: {"content": "...."}
        if isinstance(raw.get("content"), str):
            content = raw["content"]
        # 2) content come lista di pezzi
        elif isinstance(raw.get("content"), list):
            parts = []
            for p in raw["content"]:
                if isinstance(p, str):
                    parts.append(p)
                elif isinstance(p, dict) and isinstance(p.get("content"), str):
                    parts.append(p["content"])
            content = "".join(parts)
        # 3) stile OpenAI-like con "choices"
        elif isinstance(raw.get("choices"), list) and raw["choices"]:
            ch = raw["choices"][0]
            if isinstance(ch.get("text"), str):
                content = ch["text"]
            elif isinstance(ch.get("message"), dict) and isinstance(
                ch["message"].get("content"), str
            ):
                content = ch["message"]["content"]
        # Fallback: json intero
        else:
            content = text_body or str(raw)
    else:
        # Nessun JSON decente, usiamo il body testuale
        content = text_body or "[risposta LLM vuota]"

    content = (content or "").strip()
    if not content:
        content = "[LLM non ha restituito testo utile]"

    return web.json_response({"content": content})


@web.middleware
async def request_logger_middleware(request: web.Request, handler):
    start = datetime.now(timezone.utc)
    try:
        response = await handler(request)
    except web.HTTPException as exc:
        duration = (datetime.now(timezone.utc) - start).total_seconds()
        logger.warning(
            "HTTP %s %s -> %s (%.3fs) [EXC]",
            request.method,
            request.path,
            exc.status,
            duration,
        )
        raise
    except Exception as exc:  # noqa: BLE001
        duration = (datetime.now(timezone.utc) - start).total_seconds()
        logger.exception(
            "HTTP %s %s -> 500 (%.3fs) [UNHANDLED ERROR]",
            request.method,
            request.path,
            duration,
        )
        return web.json_response(
            {"error": "internal_error", "detail": str(exc)},
            status=500,
        )

    duration = (datetime.now(timezone.utc) - start).total_seconds()
    logger.info(
        "HTTP %s %s -> %s (%.3fs)",
        request.method,
        request.path,
        getattr(response, "status", "?"),
        duration,
    )
    return response


# ================== APP FACTORY ==================

def create_app() -> web.Application:
    app = web.Application(middlewares=[request_logger_middleware])

    # GUI
    app.router.add_get("/", ui_index)

    # STATIC: assets GUI (logo + textures)
    assets_path = BASE_DIR / "gui" / "assets"
    app.router.add_static("/assets/", path=str(assets_path), name="assets")

    # SYSTEM
    app.router.add_get("/system/health", health)

    # AUTH (stub locale)
    app.router.add_post("/auth/register", auth_register)
    app.router.add_post("/auth/login", auth_login)
    app.router.add_get("/auth/session/validate", auth_validate_session)
    app.router.add_get("/auth/me", auth_me)
    app.router.add_post("/auth/change_password", auth_change_password)

    # ORDERS
    app.router.add_get("/account/orders", account_orders)

    # ECOM
    app.router.add_post("/ecom/pricing", pricing)
    app.router.add_post("/ecom/order_draft", order_draft)
    app.router.add_post("/ecom/product/update", product_update)

    # ADMIN
    app.router.add_get("/admin/clients/all", admin_clients_all)
    app.router.add_post("/admin/clients/save", admin_clients_save)
    app.router.add_post("/admin/clients/delete", admin_clients_delete)
    app.router.add_post("/admin/clients/import_promo", admin_clients_import_promo)
    app.router.add_post("/admin/promo/add_points", admin_promo_add_points)
    app.router.add_get("/admin/promo/summary", admin_promo_summary)
    app.router.add_get("/admin/offers/all", admin_offers_all)
    app.router.add_post("/admin/offers/save", admin_offers_save)
    app.router.add_get("/admin/products/all", admin_products_all)
    app.router.add_post("/admin/products/save", admin_product_save)
    app.router.add_delete("/admin/products/{sku}", admin_product_delete)
    app.router.add_post("/admin/price_list/import", admin_price_list_import)
    app.router.add_get("/admin/price_list/status", admin_price_list_status)
    app.router.add_get("/admin/daily-offer", admin_daily_offer_get)
    app.router.add_post("/admin/daily-offer", admin_daily_offer_save)
    app.router.add_delete("/admin/daily-offer", admin_daily_offer_delete)

    # PUBLIC OFFER
    app.router.add_get("/offer/daily", public_daily_offer)

    # LLM
    app.router.add_post("/llm/complete", llm_complete)
    app.router.add_post("/api/llm/chat", llm_chat)

    return app


app = create_app()


def main() -> None:
    """Avvia il server API aiohttp."""
    host = os.environ.get("API_HOST", "127.0.0.1")
    port = int(os.environ.get("API_PORT", 8080))
    logger.info(
        "Avvio TheLight24 API server su %s:%s (LLM_BACKEND_URL=%s)",
        host,
        port,
        LLM_BACKEND_URL,
    )
    web.run_app(app, host="127.0.0.1", port=8080)


if __name__ == "__main__":
    main()
