#!/usr/bin/env python3
"""Importa gli ordini mensili esportati dal gestionale.

Questo script viene richiamato da ``import_orders.sh`` (cron/manuale) e legge un
file CSV/XLSX con le colonne principali dell'ordine. Configura il percorso del
file con ``--input`` o variabile ``ORDERS_INPUT_FILE``.
"""

import argparse
import csv
import os
import re
import sys
from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple

# Root del progetto TheLight24 (cartella che contiene "api", "scripts", "data", ecc.)
BASE_DIR = Path(__file__).resolve().parent.parent

# Assicura che la root del progetto sia nel PYTHONPATH così "api" è importabile
if str(BASE_DIR) not in sys.path:
    sys.path.insert(0, str(BASE_DIR))

from api.db import bulk_insert_orders, delete_orders_older_than, init_db
from core.logger import get_logger

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover - import facoltativo in ambienti minimal
    load_workbook = None

DEFAULT_INPUT = Path(
    os.environ.get("ORDERS_INPUT_FILE", BASE_DIR / "data" / "orders_latest.csv")
)

logger = get_logger("thelight24.import_orders")
EMAIL_RE = re.compile(r"[^@\s]+@[^@\s]+\.[^@\s]+")


def parse_date(value: Optional[str]) -> Optional[str]:
    """Prova a normalizzare la data dell'ordine in ISO (YYYY-MM-DD)."""

    if value is None:
        return None

    if isinstance(value, datetime):
        return value.date().isoformat()

    text = str(value).strip()
    if not text:
        return None

    patterns = [
        "%d/%m/%Y",
        "%d/%m/%Y %H:%M",
        "%Y-%m-%d",
        "%Y-%m-%d %H:%M:%S",
    ]
    for pattern in patterns:
        try:
            return datetime.strptime(text, pattern).date().isoformat()
        except ValueError:
            continue

    # come fallback tentiamo a interpretare direttamente da datetime.fromisoformat
    try:
        return datetime.fromisoformat(text).date().isoformat()
    except Exception:
        return None


def safe_number(value: Optional[str]) -> Optional[float]:
    if value is None:
        return None
    text = str(value).strip().replace(".", "").replace(",", ".")
    try:
        return float(Decimal(text))
    except (InvalidOperation, ValueError):
        return None


def load_csv_rows(path: Path) -> Iterable[Dict[str, str]]:
    with path.open(newline="", encoding="utf-8-sig") as handle:
        reader = csv.DictReader(handle)
        for row in reader:
            yield row


def load_xlsx_rows(path: Path) -> Iterable[Dict[str, str]]:
    if load_workbook is None:
        raise RuntimeError("openpyxl non disponibile nell'ambiente corrente")

    wb = load_workbook(path, read_only=True)
    ws = wb.active
    header_row = next(ws.iter_rows(min_row=1, max_row=1))
    headers = [
        (str(cell.value).strip().lower() if cell.value is not None else "")
        for cell in header_row
    ]
    for row in ws.iter_rows(min_row=2, values_only=True):
        yield {headers[i]: (row[i] if i < len(row) else None) for i in range(len(headers))}


def normalize_row(raw: Dict[str, str], row_index: int) -> Tuple[Optional[Dict[str, str]], Optional[str]]:
    def pick(*keys: str) -> str:
        for key in keys:
            if key in raw and raw.get(key) not in (None, ""):
                return str(raw.get(key)).strip()
        return ""

    document_number = pick(
        "numero documento",
        "n.doc.",
        "n.doc",
        "documento",
        "doc",
        "document_number",
    )

    status = pick("stato", "status")

    cause = pick("causale", "cause")

    customer_name = pick(
        "ragione sociale",
        "intestatario",
        "cliente",
        "customer_name",
    )

    customer_email = pick(
        "email",
        "email cliente",
        "email_cliente",
        "customer_email",
    )

    order_date = parse_date(
        pick("data ordine", "data", "order_date")
    )

    total_amount = safe_number(
        pick("tot.doc.", "tot doc", "totale", "importo", "total_amount")
    )

    external_id = pick("id gestionale", "external_id") or document_number

    notes = pick("note", "notes")

    if not document_number:
        return None, "missing_document_number"

    if not customer_email or not EMAIL_RE.match(customer_email):
        return None, "invalid_email"

    if raw.get("tot.doc.") or raw.get("total_amount"):
        if total_amount is None:
            return None, "invalid_total_amount"

    return (
        {
            "document_number": document_number,
            "status": status,
            "cause": cause,
            "customer_name": customer_name,
            "customer_email": customer_email,
            "order_date": order_date,
            "total_amount": total_amount,
            "external_id": external_id,
            "notes": notes,
        },
        None,
    )


def load_orders(path: Path) -> Tuple[List[Dict[str, str]], int, int]:
    records: List[Dict[str, str]] = []
    discarded = 0
    loader = load_csv_rows
    if path.suffix.lower() in {".xlsx", ".xlsm"}:
        loader = load_xlsx_rows

    for idx, row in enumerate(loader(path), start=2):
        try:
            normalized, error = normalize_row({(k or "").lower(): v for k, v in row.items()}, idx)
            if error or not normalized:
                logger.warning(
                    "Riga %s scartata (%s) doc=%s", idx, error or "invalid_row", row.get("document_number") or row.get("doc") or "?"
                )
                discarded += 1
                continue
            records.append(normalized)
        except Exception as exc:  # noqa: BLE001
            logger.warning("Errore parsing riga %s: %s", idx, exc, exc_info=True)
            discarded += 1
            continue
    return records, idx if 'idx' in locals() else 0, discarded


def main() -> None:
    parser = argparse.ArgumentParser(description="Import ordini mensili")
    parser.add_argument("--input", type=Path, default=DEFAULT_INPUT, help="Percorso file CSV/XLSX esportato dal gestionale")
    parser.add_argument("--retention-days", type=int, default=31, help="Giorni di storico da mantenere (default: 31)")
    args = parser.parse_args()

    input_path = Path(args.input)
    if not input_path.exists():
        print(f"[ERRORE] File di input non trovato: {input_path}")
        print("Configura il percorso nel cron o passa --input /percorso/file.xlsx")
        return

    init_db()

    removed = delete_orders_older_than(args.retention_days)
    print(f"Ordini rimossi perché più vecchi di {args.retention_days} giorni: {removed}")

    try:
        orders, total_rows, discarded = load_orders(input_path)
    except Exception as exc:  # noqa: BLE001
        print(f"[ERRORE] Impossibile leggere il file {input_path}: {exc}")
        return

    inserted = bulk_insert_orders(orders)
    valid_rows = len(orders)
    total_processed = total_rows if total_rows else valid_rows + discarded
    summary = f"Import ordini completato: tot={total_processed}, validi={valid_rows}, scartati={discarded}"
    print(summary)
    logger.info(summary)
    print(f"Ordini importati: {inserted} su {len(orders)} righe valide")
    print("Esegui questo script via cron per aggiornare ogni mese lo storico ordini.")


if __name__ == "__main__":
    main()
